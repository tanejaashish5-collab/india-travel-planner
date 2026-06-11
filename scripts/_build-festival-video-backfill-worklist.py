#!/usr/bin/env python3
"""
Build the festival B-roll backfill worklist: the festivals that still lack a
hero clip in R2. Reads the regenerated data/festivals/video-prompts.csv (501
rows for the current catalog) and apps/web/src/lib/festival-heroes.ts (the 319
slugs already in R2), and writes a focused worklist of the MISSING slugs only —
ready to feed into the AI-video generation tool.

Outputs (under data/festivals/):
  - video-prompts-missing-2026-06-03.csv   (machine-readable)
  - video-prompts-missing-2026-06-03.xlsx  (operator sheet: copy full_prompt,
    save as {slug}.mp4, tick status)

Run: python3 scripts/_build-festival-video-backfill-worklist.py
"""
import csv, re, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PROMPTS = ROOT / "data/festivals/video-prompts.csv"
REGISTRY = ROOT / "apps/web/src/lib/festival-heroes.ts"
STAMP = "2026-06-03"
OUT_CSV = ROOT / f"data/festivals/video-prompts-missing-{STAMP}.csv"
OUT_XLSX = ROOT / f"data/festivals/video-prompts-missing-{STAMP}.xlsx"

# slugs that already have a clip in R2
have = set(re.findall(r'"([a-z0-9-]+)"', REGISTRY.read_text()))
print(f"slugs with clip (registry): {len(have)}")

rows = list(csv.DictReader(PROMPTS.open(encoding="utf-8")))
print(f"prompt rows: {len(rows)}")

seen, missing = set(), []
for r in rows:
    slug = r["festival_slug"]
    if slug in have or slug in seen:
        continue
    seen.add(slug)
    missing.append(r)
print(f"MISSING (no clip yet): {len(missing)}")

COLS = ["priority_tier", "festival_slug", "festival_name", "destination_name",
        "state", "month_name", "approximate_date", "full_prompt",
        "negative_prompt", "reference_image_url", "aspect_ratio",
        "duration_seconds"]

# --- CSV ---
with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f, quoting=csv.QUOTE_ALL)
    w.writerow(COLS + ["save_as", "status"])
    for r in sorted(missing, key=lambda x: (x["priority_tier"], x["festival_slug"])):
        w.writerow([r.get(c, "") for c in COLS] + [f'{r["festival_slug"]}.mp4', "pending"])
print(f"wrote {OUT_CSV.relative_to(ROOT)}")

# --- XLSX ---
try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("openpyxl not available — CSV written, skipping xlsx", file=sys.stderr)
    sys.exit(0)

wb = Workbook()

# README sheet
readme = wb.active
readme.title = "README"
readme["A1"] = f"Festival B-roll backfill worklist — {STAMP}"
readme["A1"].font = Font(size=14, bold=True)
notes = [
    "",
    f"{len(missing)} festivals at /festivals/[slug] still have NO hero clip in R2.",
    "The hero <video> is already wired (lib/festival-heroes.ts) — these fall back to the image until a clip lands.",
    "",
    "How to use the WORKLIST sheet:",
    "  1. Copy full_prompt into the AI video tool (Veo 3.1 Lite, 16:9 — same as the prior batch).",
    "  2. Use reference_image_url as the image-to-video anchor.",
    "  3. Save the render with the EXACT name in save_as ({slug}.mp4).",
    "  4. Drop the file in data/festivals/videos/ and set status = done.",
    "",
    "When a batch is downloaded:",
    "  - node scripts/_upload-festival-videos.mjs        (uploads new clips to R2)",
    "  - node scripts/_verify-festival-video-sync.mjs    (byte-verify before deleting local)",
    "  - node --env-file=apps/web/.env.local scripts/_gen-festival-hero-registry.mjs   (refresh registry)",
    "  - commit the regenerated apps/web/src/lib/festival-heroes.ts + redeploy",
    "",
    "Guardrail: prompts are LOCATION-ONLY B-roll (the place at festival time), never the ritual/people/attire.",
]
for i, line in enumerate(notes, start=3):
    readme[f"A{i}"] = line
readme.column_dimensions["A"].width = 110

# Worklist sheet
ws = wb.create_sheet("WORKLIST")
headers = COLS + ["save_as", "status"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="1F2937")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = hfill
    c.alignment = Alignment(vertical="center")

for r in sorted(missing, key=lambda x: (x["priority_tier"], x["festival_slug"])):
    ws.append([r.get(c, "") for c in COLS] + [f'{r["festival_slug"]}.mp4', "pending"])

# status dropdown
dv = DataValidation(type="list", formula1='"pending,done,skip"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"N2:N{ws.max_row}")

widths = {"A": 12, "B": 30, "C": 30, "D": 22, "E": 16, "F": 12, "G": 16,
          "H": 70, "I": 40, "J": 40, "K": 12, "L": 10, "M": 30, "N": 10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

wb.save(OUT_XLSX)
print(f"wrote {OUT_XLSX.relative_to(ROOT)}")

# tier breakdown
from collections import Counter
tc = Counter(r["priority_tier"] for r in missing)
print("by tier:", dict(sorted(tc.items())))
