#!/usr/bin/env python3
"""Build AI-build-ideas FULL deliverable (xlsx + csv) — V2, full-depth.
Merges 12 channel extracts (v1) + 6 new channels + 5 newsletters + demand DBs.
First-principles, deduped, tiered, scored. Adds cross-source CONSENSUS count,
a demand-validated real-revenue proof table, and meta-playbooks.
Run: python3 scripts/_build-ai-build-ideas-v2.py
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.join(os.path.dirname(__file__), "..", "data", "ideas")
os.makedirs(OUT, exist_ok=True)
DATE = "2026-06-02"

# ------- MASTER IDEAS -------
# Rank, Tier, Idea, One-liner thesis, Moat type, Moat strength, #Sources, Build difficulty, Non-tech fit, Score, Verdict, Note
M_COLS = ["Rank","Tier","Idea","One-line thesis","Moat type","Moat","#Sources","Build difficulty","Non-tech-founder fit","Score /5","Verdict","Why defensible — or why it's a trap"]
M = [
 # ---------- TIER A ----------
 [1,"A","Regulatory-arbitrage compliance SaaS","Pick a painful, boring compliance niche with QUANTIFIABLE $ savings (expat state-tax domicile, sales-tax nexus, licensing) and own the workflow + rules.","Regulatory knowledge + quantifiable ROI","Strong","2","High for a non-expert (domain + liability)","NO — needs licensed CPA/lawyer co-founder","3","AVOID (validated 2026-06-02)","VALIDATION KILLED THIS for a non-expert solo founder. Savvy Nomad revenue is ~$360K/yr (the '$1.7M' was unverified) AND it has a CPA co-founder — not the non-tech win it looked like. Zero documented non-domain solo compliance-SaaS successes post-2020. Real UPL liability (Upsolve case). Only viable with a licensed co-founder (50% split, 30-40% margins, 18-24mo). See data/ideas/_validation/."],
 [2,"A","Vertical AI Operator (own ONE niche)","Own one high-value vertical (e-comm returns, insurance claims, clinic front-desk, accounting) end-to-end — sell the OUTCOME, not 'an AI'.","Domain data + ROI playbooks + integrations","Strong","6","Medium","PARTIAL — needs domain depth/partner","5","BUILD","Strongest cross-source consensus (Nate/Devin $4-5M ARR, YC Replit, demand-DBs, Matthew Berman, Neuron). Moat = vertical data + case studies, not the model. 'AI compresses every moat' EXCEPT domain+regulatory."],
 [3,"A","Winning-Ad Deconstructor","Monitor competitor ad libraries, auto-score winners, break each down frame-by-frame (hook, pacing, copy, music, CTA) into a repeatable creative spec.","Data (ad-performance corpus that compounds)","Strong","2","Medium (vision+scraping)","YES — marketer buyer","5","BUILD","Real paying market (Foreplay/Atria $99-300/mo). Corpus compounds. Risk: Meta fights library scraping."],
 [4,"A","Agent Governance / Observability / Audit","The compliance-grade monitoring + audit-trail + rollback layer for companies running fleets of production agents.","Failure-mode data + compliance/audit posture","Medium-Strong","7","Medium-High","NO — ML-ops/compliance buyer","4","EXPLORE","TOP cross-source consensus (TLDR, Ben's, Neuron, demand-DBs, YC, Greg, Cole). 'Grok killed a town in 4 days'; only 21% of enterprises have oversight. Differentiate on GOVERNANCE/audit (regulatory), not generic dashboards (Datadog/LangSmith encroaching)."],
 [5,"A","Voice AI agents for customer service (vertical)","LLM phone agents that resolve 60-95% of inbound/outbound calls end-to-end in ONE industry, human handoff on edge cases.","Industry playbooks + carrier relationships + compliance","Medium-Strong","3","Medium-High","PARTIAL","4","EXPLORE","Proven ROI (2-3mo payback), demand-DBs 4/5, GigaML beat funded Sierra on execution. Niche to one vertical; carrier + compliance = moat. Crowded but real."],
 [6,"A","AI Sales-Call Sparring Partner","Voice AI that role-plays buyer objections, coaches reps live, scores on a rubric, tracks improvement — per-vertical objection libraries.","Objection library + outcomes data","Medium-Strong","1","Medium","YES — sales-leader buyer","4","BUILD","Incumbent rep training $5K+/rep. Async practice is the wedge; integrations sticky."],
 [7,"A","Dark Factory — gated autonomous shipping","Agents that ship code unattended: issue→PR→multi-agent review→tests→canary→monitor→auto-promote/rollback; learns from deploy metrics.","Private deploy-outcome metrics (compounding)","Strong","2","Very High (months + prod infra)","NO — deep technical","4","EXPLORE","Highest ceiling + strongest moat, but hardest build; one bad deploy breaks prod. Team play."],

 # ---------- TIER B ----------
 [8,"B","Niche app + single-influencer distribution","Not one idea, a PLAYBOOK: niche pain → ship in 30 days with AI → ONE aligned influencer on profit-share → hard paywall.","Distribution lock-in + brand + retention","Medium","4","Low-Medium","YES — this is the non-tech path","4","BUILD","DEMAND-VALIDATED 4×: Cut Coach $20K/mo, Money AI $35K/mo, Locked $14K/mo (14yo founder), Glam Mob $150-250K/mo. The product is almost irrelevant; the ONE-influencer loop is the 10x lever. See PLAYBOOKS tab."],
 [9,"B","Condition-specific Health 'Brain'","One condition only (GERD/migraine/IBS): aggregate bloodwork+biome+diet+specialist into a decision layer; monetize plans, supplements, matching.","Medical-data lock-in + regulatory + LTV","Strong","2","Medium-High (regulatory)","PARTIAL — clinical credibility","4","EXPLORE","Very high WTP ($200-2K/yr), huge TAM. demand-DBs lists AI personalized medicine 4/5 ($100B+). Regulatory complexity is the cost of the moat."],
 [10,"B","IRL community / anti-loneliness membership","Hybrid online+offline membership pairing people for real experiences; revenue from events + membership.","Network effects + offline switching costs","Medium-Strong","1","Low-Medium (ops-heavy)","YES — barely any code","4","BUILD","Most durable, least-AI idea — won't be eaten by a model release. Ops-intensive not eng-intensive. Strong non-technical fit. (Proven retreat economics $90K/event.)"],
 [11,"B","Niche language learning (BeyondEnglish)","Quality courses for underserved languages (Basque, Quechua, Yiddish) with native instructors; diaspora/grant-funded GTM.","Curriculum + instructor network + community","Medium-Strong","1","Medium","YES","4","EXPLORE","Greg's top pick. Duolingo ignores the long tail; community lock-in + high margin; founder-buildable LMS+Stripe."],
 [12,"B","Niche community / hobby studios","Pick ONE intense community (golfers, tabletop, streamers); ship 3 products in 90 days; monetize via the community.","Community relationships + insider access","Medium","1","Low-Medium","YES","4","EXPLORE","Greg pattern. Defensible via insider access; recurring revenue; founder-buildable. Distribution is the moat."],
 [13,"B","Agent-first 'Action App' for a vertical","Mobile app that DOES the task end-to-end (books, cancels, files, disputes) instead of chatting — built agent-first.","First-mover vertical UX + workflow depth","Medium","1","Medium","PARTIAL","4","EXPLORE","Greenfield vs legacy click-based apps incumbents can't rebuild without cannibalizing. Churns if unreliable."],
 [14,"B","Event / moment capture app","Time-bound shared-album app for weddings/parties (digital disposable camera); constraint-driven UX as the feature.","Network effects (per event) + brand","Weak-Medium","1","Medium","YES","3","EXPLORE","DEMAND-VALIDATED: 'Once' hit $20K/mo in 83 days. Weak long-term moat (copyable); retention is the problem. Notable for its PRE-VALIDATION rigor (tested at a party before building)."],
 [15,"B","Elder-tech for the 65+ market","AI hearing/vision/memory/social aids for a cash-rich, founder-ignored demographic.","Demographic trust + regulatory","Medium","1","Medium","PARTIAL","3","EXPLORE","Large TAM, low competition; different acquisition channels; trust-first slow market."],
 [16,"B","AI-native niche media → owned product","Build a high-quality (NOT slop) audience channel in one vertical; audience = owned acquisition channel for a product.","Audience trust (platform-independent)","Medium","1","Low-Medium","YES","3","EXPLORE","Audience is the moat only if quality is high; slop dies fast. Channel = funnel, not product."],
 [17,"B","Pet-health smart monitoring","IoT + AI for pet health (sleep, HR, anomalies); highest-WTP demographic, <2% penetration.","Demographic WTP + community","Medium","1","Medium (hardware+sw)","PARTIAL","3","EXPLORE","Real demand, low competition, but hardware+software combo."],

 # ---------- TIER C ----------
 [18,"C","Personal software generator (SaaS unbundling)","Describe a workflow in English → agents generate a custom tool from your SaaS APIs (Gmail/HubSpot/Stripe). Replace 50 SaaS subs.","Integration template library + generation quality","Medium","2","Medium","PARTIAL","4","EXPLORE","Ben's Bites: founder replaced a $40/mo tool in 2h with Codex. Strong psychological pull (unbundle SaaS). Narrow to one domain first."],
 [19,"C","Agent data governance / PII audit","Track data provenance through multi-agent workflows; flag PII exposure; enforce GDPR/HIPAA at each agent step.","PII-detection models + compliance integrations","Medium-Strong","2","Medium (+legal)","NO — compliance buyer","4","EXPLORE","TLDR #10 + rising fraud (187% AI-traffic spike). Compliance = regulatory moat. Needs legal/security depth."],
 [20,"C","Adversarial agent testing harness","Ensemble (fuzzer/security/perf/chaos) that auto-tests agent-written code → test suite + vuln report.","Agent ensemble + feedback loop","Medium","1","Medium-High","NO — technical","4","EXPLORE","Real leverage as vibe-coded volume explodes; differentiate via bug-bounty integration."],
 [21,"C","'Dark-code' QA / validation for production","Validate AI-generated systems before prod: success-criteria, edge-case tests, audit of code nobody read.","Testing library + checklist benchmark","Medium","3","Medium","PARTIAL","4","EXPLORE","Consensus (Nate, TLDR, Ben's). Genuine 'dark code' gap; moat softens ~18mo as validation bakes into models — move fast or niche to regulated vertical."],
 [22,"C","Agent auto-validation loop","Drop-in CI: fresh-context reviewer agents + rule engine (build/test/secret/perf) + auto-fix + learning KB.","Proprietary fix-outcome KB","Medium","1","Medium","NO — technical","4","EXPLORE","High leverage; plugs into existing CI so low adoption friction."],
 [23,"C","Spatial video from a sketch","Hand-drawn camera path on an image → 8-10s drone-POV video with correct 3D + foley. For real-estate/architecture/game-preview.","Domain data (if you own it) + speed","Medium (temporary)","1","Medium-High","NO","3","EXPLORE","Matt Wolfe: Gemini Omni does this now. ~6-month window before Sora/Claude ship parity. Only with a fast vertical wedge + owned data."],
 [24,"C","Sovereign / open-model fine-tuning service","Hosted fine-tune of open weights (Cohere Command A+, Llama) for a vertical; run frontier-class locally, no API lock-in.","Domain datasets + pipeline tooling + trust","Medium","3","Medium","PARTIAL","3","EXPLORE","Consensus (Neuron, TLDR, demand-DBs). Real 'sovereign AI' pull, but together.ai/HF/Replicate exist + consolidation likely by 2027. Win on a vertical."],
 [25,"C","Self-evolving agent knowledge base","Institutional-memory agents read+write decision logs/failures so the next agent skips repeated mistakes.","Per-codebase proprietary KB","Medium","4","Medium","NO","3","EXPLORE","Consensus (Cole, Nate, Jack, YC). Infra commoditized (Supabase+pgvector); likely a feature, not a business."],
 [26,"C","AI-to-human handoff orchestration","Long-horizon agents (35h runs) escalate the few decisions needing human approval, then resume — without killing autonomy.","Escalation UX + task-queue integrations","Medium","1","Medium","PARTIAL","3","EXPLORE","Neuron: Qwen ran 35h, hallucinated mid-run. No standard escalation protocol; every company rebuilds it. Fragmented TAM."],
 [27,"C","Agent-to-phone bridge","Mobile-first control of long-running desktop agents: check status, feed input, approve, cancel via push.","First-mover UX + platform integrations","Weak-Medium","1","Medium","PARTIAL","3","EXPLORE","Ben's Bites recurring ask. Real gap; start as web + deep-link from iMessage/Telegram. Platforms may ship native."],
 [28,"C","Vision module for headless agents","Let CLI/API agents 'see' dashboards/screenshots → structured JSON per frame.","Frame-parsing taxonomy + temporal consistency","Weak-Medium","1","Medium","NO","3","EXPLORE","Useful wedge; high risk Claude/Hermes ship native agent-vision in 3-6mo."],
 [29,"C","Agent skill marketplace","Versioned shareable agent skills with usage analytics + royalty splits to authors; cross-platform sync.","Marketplace network effects","Weak-Medium","2","Medium-High","NO","3","EXPLORE","Greg + Ben's. Dead until 10K+ active users on a platform you don't control; vendors may own this layer."],

 # ---------- TIER D (HYPED TRAPS) ----------
 [30,"D","Multi-model cost router / spend optimizer","Route each task to the cheapest sufficient model + dashboard savings.","Routing heuristics","WEAK","8","Low-Medium","NO","2","AVOID","THE trap. 8+ sources pitch it; OpenRouter just raised $1.3B (incumbent won). Consensus here = sell signal, not buy. Anthropic ships native."],
 [31,"D","Multi-agent orchestration framework","Generic framework to coordinate teams of agents (researcher+writer+reviewer).","Evals + observability","WEAK","6","Medium-High","NO","2","AVOID","LangGraph/CrewAI/AutoGen/Anthropic own this. Debugging multi-agent is hard, ROI vs single-agent unclear. Infra, not a business."],
 [32,"D","Generic 'Company Brain' knowledge layer","Unified KG pulling Slack/email/docs/DBs for agents.","Connector breadth","WEAK","3","Medium","Partial","2","AVOID","'Anyone can build a denormalized table' (YC's own words). Becomes a feature of an agent platform; connector arms race."],
 [33,"D","AI-native SaaS replacement ('SaaS challengers')","Rebuild Salesforce/Zendesk AI-native + cheaper.","2-yr lock-in (weak)","WEAK","2","Medium-High","Partial","2","AVOID","Low switching cost + pricing wars. Incumbents add AI faster than you build distribution."],
 [34,"D","Model drift / honesty canary","Daily canaries that alert when a provider silently changes model behavior.","Canary dataset","Weak-Medium","1","Low-Medium","NO","2","EXPLORE","More defensible than the rest of D (dataset is yours) but 'who pays for a warning?' unsolved. Validate WTP first."],
 [35,"D","Benchmark aggregator / eval consolidator","Run a model across 20+ public benchmarks → capability card.","Aggregation only","WEAK","2","Medium","NO","1","AVOID","Aggregating public data = no moat. Benchmark owners would rather own the platform."],
 [36,"D","Cross-agent config abstraction","Normalize CLAUDE.md/agents.md/.codex so projects port between frameworks.","Spec + integrations","WEAK","2","Low","NO","2","AVOID","Doomed by vendors shipping their own normalization."],
 [37,"D","Local fine-tuning / 'personal AI twin'","One-click local fine-tune on your data/values; sovereignty narrative.","Brand trust","WEAK","2","Medium","Partial","2","AVOID","Slim margins, tiny market, open-source algos. Narrative, not moat. (Distinct from the VERTICAL fine-tune service #24, which is more defensible.)"],
 [38,"D","Agent-native terminal (Linux/Windows)","Modern multi-agent terminal UI for devs stuck on Tmux.","Workflow integration","WEAK","1","High (OS rendering)","NO","2","AVOID","~$1.5M TAM, commodity rendering. Dev love, no money."],
 [39,"D","Generic no-code agent builder / enablement","No-code platform for non-technical teams to build agents.","Vertical templates","WEAK","2","Medium","Partial","2","AVOID","Make/Zapier/Cowork already moving up-market. Only a vertical-template wedge survives — which is really idea #2."],
 [40,"D","Repository intelligence / AI code review","Agent that reviews PRs, refactors, catches security issues.","GitHub integration","WEAK","3","Medium","NO","2","AVOID","Cursor/Copilot/CodeRabbit ship this; GitHub adds it natively. No uncontested segment."],
 [41,"D","Benchmark integrity / synthetic-task gen","Contamination-free domain benchmarks / synthetic training tasks.","Verifier design","WEAK","2","Medium","NO","2","AVOID","Research-tier, hard to monetize; labs rebuild it. 'Who pays?' unsolved."],
 [42,"D","Prompt-caching optimizer dashboard","Maximize prompt-cache hits, cut token spend.","Analytics + templates","WEAK","2","Low","Partial","2","AVOID","~12-month runway then Anthropic ships native. Free lead-gen tool at best."],
 [43,"D","Thinking-time budget optimizer","Auto-tune model effort/thinking-time vs task difficulty.","Classifier","WEAK","2","Low","NO","2","AVOID","A setting, not a startup. Native-feature risk <3mo."],
 [44,"D","Capital-heavy frontier plays (chips/robotics/biotech)","Inference chips, humanoid sim-to-real, robotics training-data pipelines, AI drug-discovery.","Patents + hardware + data","Strong","3","Very High (2-5yr, $5M+)","NO","2","AVOID (for you)","Real TAM and strong moats, but 2-5 years + $5M+ + teams. Listed so you know they're NOT solo/non-technical plays. Watch as trends, don't build."],
]

# ------- DEMAND-VALIDATED REAL BUSINESSES (proof points) -------
DV_COLS = ["Business","Revenue (stated)","What it is","Moat","Solo / non-tech?","Key takeaway"]
DV = [
 ["Savvy Nomad","~$360K/yr ($30K/mo verified; '$1.7M' UNVERIFIED)","Expat US state-tax domicile SaaS","Regulatory + quantifiable ROI","NO — has a CPA co-founder (Jameson) + Bohdan","CORRECTION (2026-06-02 validation): NOT a non-technical solo win — required a licensed CPA from day one. Quantifiable-ROI thesis holds; the 'anyone can build it' framing does not."],
 ["Perkeep Chat","$1.5M lifetime / $300K-yr","AI comms + legal research for incarcerated users","Monopoly (regulated, closed ecosystem)","Yes","Closed ecosystem = competitors can't acquire users externally. Extreme demand validation (20% TAM)."],
 ["Task Magic","$3M ARR → 7-fig exit","No-code browser RPA (Zapier-limit users)","Tentpole ecosystem + SEO","Yes (1 CTO hire)","Subsidiary products (MailLead/LeadQuest) funnel to core. Ecosystem lock-in beats feature parity."],
 ["Shipyard","$307K ARR","AI app builder for 100% non-technical","Niche differentiation (12-18mo)","Partial","Read 100% of 5 competitors' reviews → built the #1 complaint (mobile). Niche down on a weakness."],
 ["Glam Mob + Sprout","$150-250K/mo each","Makeup glow-up + job-prep apps","Systematized UGC creator network","No (co-founder + ops)","200+ trained micro-creators, 400-500M views/mo. App is irrelevant; distribution system IS the business."],
 ["Money AI","$300/mo → $35K/mo","Minimalist AI expense tracker","Single-influencer distribution","Yes","ONE Colombian influencer = 100x MRR in a month. Key-person risk, but proves the one-influencer lever."],
 ["Cut Coach","$20K/mo","Wrestling weight-cut protocol app","Niche distribution + seasonal","Yes","Founder's own pain (wrestler). $30 to avoid $100 failure cost. Built in 1 month with Cursor."],
 ["Locked","$14K/mo","Gamified habit/fitness app (14yo founder)","ASO + influencer exclusivity","Yes (w/ brother)","One influencer video = 1M views, 1,800 installs, $3K. CPM < RPM = profitable per video."],
 ["Once","$20-22K/mo (83 days)","Digital disposable camera for events","Network effects per event (weak)","Yes","PRE-VALIDATED at a friend's party BEFORE writing the mobile app. Rigor most founders skip."],
 ["Snag","$30K/mo","Free-items-near-you marketplace","Distribution + app-store rank","Yes","Reverse-engineered from Sensor Tower ('what apps make money?') → built a 10% better UX."],
 ["Replit","$2.5M → $250M/yr","Agent coding platform","Market creation (weak long-term)","No","Cheap software creation unlocks new models. Market-creation beats zero-sum capture (but copyable)."],
 ["Razorpay","India fintech giant","Payments / lending","Regulatory (RBI approval gauntlet)","No","12-month approval kills 99% of competitors. Regulation flips from bug to feature. (YC: 'AI compresses every moat' — except this.)"],
]

# ------- PLAYBOOKS / META-PATTERNS -------
PB = [
 ["1. Single-aligned-influencer distribution","The 10x lever across 4+ demand-validated apps. Find a niche pain you have → ship MVP in 30 days with Cursor/Claude → find ONE influencer whose audience IS your niche → profit-share or flat CPM < your RPM → hard paywall (3-day trial). NOT 100 influencers — one hand-crafted relationship, 3-5 videos/mo. Niche audiences convert 5-10% vs 0.1% on ads.","Cut Coach, Money AI, Locked, Glam Mob"],
 ["2. Regulatory / compliance arbitrage","Boring, unsexy, PAINFUL niche where the customer can compute exact $ saved. Regulatory knowledge becomes the moat (slow for competitors, trusted by customers). Quantifiable ROI ('save $2,400, costs $500') kills the sales objection.","Savvy Nomad ($1.7M/yr), Perkeep Chat, Razorpay"],
 ["3. Reverse-engineer a proven market, niche on a weakness","Don't invent. Find a space already making money (Sensor Tower / Trustpilot / competitor Discords), read 100% of competitor reviews, build the #1 complaint. Position for the segment incumbents ignore (e.g., '100% non-technical').","Shipyard, Snag, Peptide AI"],
 ["4. Pre-validate with a commitment metric BEFORE building","Get real usage on a 1-week web MVP (or a printed invite at a party) before writing the full app. Eliminates builder's-block risk; launch with confidence.","Once ($20K/mo in 83 days)"],
 ["5. Vertical/data/community depth > horizontal engineering","'AI is compressing every moat' (YC/Razorpay). The only moats left: proprietary DATA, REGULATORY barriers, DISTRIBUTION/audience, COMMUNITY. NOT model access, NOT a clever wrapper, NOT raw engineering — those commoditize in 3-18 months.","YC (Harshil Mathur), Nate/Devin, Cole, every newsletter"],
 ["6. Consensus != moat (the trap detector)","The MORE sources pitch an idea, the more likely it's a commodity race. Cost-routers, multi-agent frameworks, eval aggregators, code-review bots appear everywhere AND are the weakest — incumbents already won or platforms ship native. Treat unanimous hype as a SELL signal.","Cost router (8+ sources, OpenRouter raised $1.3B)"],
]

def styled_header(ws, cols):
    fill = PatternFill("solid", fgColor="1F2937")
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

def build():
    wb = Workbook()
    # README
    ws0 = wb.active; ws0.title = "README"
    readme = [
        ["NakshIQ — AI 'What to Build' Idea Mine (FULL / V2)",""],
        ["Generated", DATE],
        ["Scope", "12 channels (v1) + 6 new channels + 5 newsletters + YC RFS/Product Hunt/Trends. ~150 videos + ~40 newsletter issues. All Haiku extraction, first-principles, NO existing-business mapping."],
        ["Channels", "Greg Isenberg, Nate Herk, Cole Medin, Jack Roberts, AI Explained, David Ondrej, My First Million, Y Combinator, Starter Story, Matt Wolfe, Matthew Berman, (Riley Brown + Liam Ottley = 0 ideas, all tutorials/AAA)"],
        ["Newsletters", "TLDR AI, Ben's Bites, The Neuron, Greg's Late Checkout, + YC RFS / Product Hunt / MIT+MSFT trend lists. (The Rundown AI = Cloudflare-blocked.)"],
        ["Raw extracts", "data/ideas/_raw/*.md  |  Transcripts: .scrapes/youtube/yt-<id>/"],
        ["",""],
        ["SHEETS",""],
        ["MASTER (ranked)", "44 deduped ideas, tiered A-D, scored, with #Sources = cross-source consensus count."],
        ["DEMAND-VALIDATED", "Real businesses with stated real revenue — proof templates, not hypotheses."],
        ["PLAYBOOKS", "6 meta-patterns that recur across the demand-validated wins. Read these FIRST."],
        ["",""],
        ["TIERS",""],
        ["A","Defensible + real money + non-obvious. The shortlist."],
        ["B","Durable consumer/market plays (least likely to be eaten by a model release)."],
        ["C","Defensible-ish dev tooling (technical buyer, closing moat windows)."],
        ["D","HYPED TRAPS — listed so you AVOID them. The most-pitched ideas live here."],
        ["",""],
        ["#Sources","How many independent sources surfaced the idea. HIGH consensus on a Tier-D idea = trap. HIGH consensus on Tier-A (e.g. Vertical Operator=6) = real."],
        ["Verdict","BUILD = worth a validation sprint now | EXPLORE = real but caveated | AVOID = looks sexy, isn't defensible."],
        ["",""],
        ["TOP 3 FOR A NON-TECHNICAL FOUNDER","#1 Regulatory-arbitrage SaaS (Savvy Nomad proof) · #3 Winning-Ad Deconstructor (marketer buyer) · #8 Niche-app + single-influencer playbook (4× demand-validated)."],
        ["BIGGEST TRAP","#30 cost-router: 8+ sources pitch it, OpenRouter raised $1.3B. Consensus = sell signal."],
        ["NEXT STEP","Competitive-landscape 'prove it' validation on ONE Tier-A pick before building."],
    ]
    for r in readme: ws0.append(r)
    ws0.column_dimensions["A"].width = 26; ws0.column_dimensions["B"].width = 118
    ws0["A1"].font = Font(size=14, bold=True)
    for i in range(1, len(readme)+1):
        ws0.cell(row=i, column=1).font = Font(bold=True)
        ws0.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # MASTER
    ws = wb.create_sheet("MASTER (ranked)")
    styled_header(ws, M_COLS)
    tier_fill = {"A":"D1FAE5","B":"DBEAFE","C":"FEF3C7","D":"FEE2E2"}
    for row in M: ws.append(row)
    for r in range(2, len(M)+2):
        fill = PatternFill("solid", fgColor=tier_fill.get(ws.cell(row=r,column=2).value,"FFFFFF"))
        for c in range(1, len(M_COLS)+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.fill = fill
    for col,w in zip("ABCDEFGHIJKL",[5,4,26,44,24,9,8,16,18,7,9,60]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # DEMAND-VALIDATED
    ws2 = wb.create_sheet("DEMAND-VALIDATED")
    styled_header(ws2, DV_COLS)
    for row in DV: ws2.append(row)
    for r in range(2, len(DV)+2):
        for c in range(1, len(DV_COLS)+1):
            ws2.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    for col,w in zip("ABCDEF",[18,18,30,26,16,60]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # PLAYBOOKS
    ws3 = wb.create_sheet("PLAYBOOKS")
    styled_header(ws3, ["Pattern","How it works","Evidence"])
    for row in PB: ws3.append(row)
    for r in range(2, len(PB)+2):
        for c in range(1,4):
            ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    for col,w in zip("ABC",[34,96,34]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    out = os.path.join(OUT, f"ai-build-ideas-FULL-{DATE}.xlsx")
    wb.save(out)
    # CSV of master
    cout = os.path.join(OUT, f"ai-build-ideas-FULL-{DATE}.csv")
    with open(cout, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(M_COLS); w.writerows(M)
    return out, cout

if __name__ == "__main__":
    x, c = build()
    print("XLSX:", os.path.abspath(x))
    print("CSV :", os.path.abspath(c))
    print(f"Master ideas: {len(M)} | A:{sum(1 for r in M if r[1]=='A')} B:{sum(1 for r in M if r[1]=='B')} C:{sum(1 for r in M if r[1]=='C')} D:{sum(1 for r in M if r[1]=='D')}")
    print(f"Demand-validated: {len(DV)} | Playbooks: {len(PB)}")
