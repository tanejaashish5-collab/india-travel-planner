#!/usr/bin/env python3
"""Validated idea map — every idea + web-researched verdict (2026-06-02).
KILL = served/commoditized/capital-or-credential-gated. CONDITIONAL = alive only with a named unlock.
SURVIVE (clean) = passes all 4 gates for a lean non-technical founder. (Result: 0.)
Run: python3 scripts/_build-validated-map.py
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.join(os.path.dirname(__file__), "..", "data", "ideas")
DATE = "2026-06-02"
COLS = ["Idea","Tier","Verdict","Killed by / incumbent (funding)","If CONDITIONAL — the unlock you'd need","Non-tech solo fit"]

# Verdict, killer, unlock
R = [
 ["Regulatory-arbitrage compliance SaaS","A","KILL","Vanta/Drata/Avalara consolidating; Savvy Nomad ~$360K + CPA co-founder","—","No — needs licensed co-founder + liability (UPL/Upsolve)"],
 ["Vertical AI Operator (own one niche)","A","KILL","Loop Returns $176M; dental 15 platforms; insurance enterprise+HIPAA","—","No — verticals already funded; integrations/data moat"],
 ["Winning-Ad Deconstructor","A","KILL","Motion $42M, Atria, Foreplay ($1.2M ARR), MagicBrief→Canva, DatAds","—","No — frame-by-frame spec already shipped by 5 funded tools"],
 ["Agent governance / observability","A","KILL","Langfuse/Braintrust/Galileo/Arize/LangSmith ($80M-$15B)","—","No — 'governance' is a bolt-on feature, not a wedge"],
 ["Voice AI agents for customer service","A","KILL","Sierra, Decagon, Vapi $500M, Bland $65M, Retell","—","No — funded knife-fight; vertical copied in 4 weeks"],
 ["AI Sales-Call Sparring Partner","A","KILL","Hyperbound $15M, Second Nature $22M, Quantified","—","No — incumbents own data + Zoom distribution"],
 ["Dark Factory — gated autonomous shipping","A","KILL","Devin/Cognition $10.2B, Factory.ai $1.5B, Copilot Workspace","—","No — owned by funded labs; GitHub bundles free"],
 ["Niche-app + single-influencer playbook","B","CONDITIONAL","Trend apps exist; example apps (Cut Coach etc.) unverifiable = survivorship-bias risk","A PRE-EXISTING aligned 50K-500K niche influencer + disciplined A/B testing (not '30-day launch'). Apple ASO 2026 penalizes novelty.","Maybe — IF you secure the influencer. The only non-capital 'new build' path."],
 ["Condition-specific Health 'Brain'","B","KILL","Midi $1B, Allara $38M, Oshi/Cylinder/Function/Zoe","(GERD-only had no competitors AND no demand signal)","No — needs clinician staff; diagnostic liability"],
 ["IRL community / anti-loneliness","B","KILL","Meetup (50M users); 222 raised $1.45M and stalled","—","No — network effects are LOCAL; per-city ops kills economics"],
 ["Niche language learning (BeyondEnglish)","B","KILL","italki ($500M+ rev, tutors for Basque/Quechua already)","—","No — long tail is unmonetizable (<500K DAU/lang); CAC>LTV"],
 ["Niche community / hobby studio","B","CONDITIONAL","Skool/Circle/Mighty own the platform layer","Reframe as a FOUNDER-LED content studio in a niche you ALREADY have audience/expertise in (e.g. travel). Then it's distribution, not tech.","Yes — IF the niche is one you already own an audience in"],
 ["Agent-first action app","B","KILL","Pine AI (53.7K users, $3M recovered)","—","No — FTC click-to-cancel + financial liability"],
 ["Event / moment capture app","B","KILL","Lapse $42M ($150M val, 100M photos/mo) + 8 clones","—","No — low LTV (1-2 events/yr), CAC>LTV, trend plateau"],
 ["Elder-tech for 65+","B","KILL","ElliQ/Intuition Robotics, GrandPad, Cera $1B","—","No — hardware-gated + Medicaid regulatory"],
 ["AI-native niche media → owned product","B","CONDITIONAL","Generic 'audience channel' = a blog, not a business","A SPECIFIC workflow-SaaS wedge + existing credibility + presale validation <$50K. Audience must already exist.","Yes — IF you already have the audience + a concrete product wedge"],
 ["Pet-health smart monitoring","B","KILL","Tractive (1.4M users); Whistle shutting down Aug-25 = consolidation","—","No — $250-500K hardware R&D + FCC + inventory"],
 ["Personal software generator / SaaS unbundling","C","KILL","Lovable $400M ARR, Replit $9B, Cursor $2B ARR, Bolt/v0","—","No — fastest-growing software category; zero solo wedge"],
 ["Agent data governance / PII audit","C","CONDITIONAL","Skyflow (agent PII, Dec-25), Credal, BigID, OneTrust","A paying healthcare/fintech customer already on Skyflow who needs the AUDIT layer on top. Contract-first, $100K+.","No — ML + compliance + cloud partnerships"],
 ["Adversarial agent testing harness","C","KILL","Antithesis $77M; Claude Code/Copilot ship native","—","No — platform-killable, skeptical dev buyer"],
 ["'Dark-code' QA / validation for prod","C","CONDITIONAL","Snyk/GitHub will ship H2-26; Graphite/Greptile ahead","Production-traffic DATA partnerships + compliance-audit positioning (not dev velocity). Needs telemetry access.","No — sales + data partnerships you can't run solo"],
 ["Agent auto-validation loop","C","KILL","Anthropic Claude Code + Copilot Workspace (free/bundled)","—","No — native to the platforms"],
 ["Spatial video from a sketch","C","KILL","Runway, Google Veo/Flow, Kling, Luma (ship camera control)","—","No — feature-ized by video-model incumbents weekly"],
 ["Sovereign / open-model fine-tuning service","C","KILL","Together.ai, Fireworks ($0.48/1M), Predibase, OpenPipe (acq.)","—","No — commoditized; $250K can't build infra moat"],
 ["Self-evolving agent knowledge base","C","CONDITIONAL","Mem0 $24M, Letta, Zep, Cognee, Supermemory","Narrow to a REGULATED vertical (healthcare/legal/finance) selling compliance audit, not horizontal memory.","No — needs domain compliance expertise"],
 ["AI-to-human handoff orchestration","C","CONDITIONAL","LangGraph, Temporal, HumanLayer own the generic layer","An explicit vertical + PRE-COMMITTED customers before building.","No — framework-native; feature not business"],
 ["Agent-to-phone bridge","C","KILL","Cursor, Google, Vercel ship native; Omnara/Conductor OSS","—","No — platform-killable within 12 months"],
 ["Vision module for headless agents","C","KILL","Claude Computer Use, OpenAI Operator, Browser Use (78K stars)","—","No — native model feature"],
 ["Agent skill marketplace","C","KILL","Anthropic official marketplace (May-26, 600 skills day 1)","—","No — platform-owned; reverse network effect"],
 ["Multi-model cost router / spend optimizer","D","KILL","OpenRouter $1.3B ($113M Series B, 25T tok/wk), Portkey, LiteLLM","—","No — the #1 trap: 8+ sources, incumbent already won"],
 ["Multi-agent orchestration framework","D","KILL","LangGraph $125M, CrewAI, OpenAI Agents SDK","—","No — consolidated in 9 months"],
 ["Generic 'Company Brain' knowledge layer","D","KILL","Glean $7.2B/$300M ARR, Hebbia $700M, MS Copilot","—","No — won 18 months ago"],
 ["AI-native SaaS replacement ('SaaS challengers')","D","CONDITIONAL","Lightfield $300M proved wedge; CRM saturated","Net-new teams + a SINGLE vertical only. Horizontal = dead.","No — vertical sales motion + funding"],
 ["Model drift / honesty canary","D","CONDITIONAL","Langfuse/Helicone (YC) + Freeplay bundle drift","2-3x better/cheaper than Langfuse scorers OR a finance/healthcare compliance vertical.","No — 'who pays for a warning?' unsolved"],
 ["Benchmark aggregator / eval consolidator","D","KILL","Artificial Analysis, LMArena, HF leaderboards (free)","—","No — no WTP for a capability card"],
 ["Cross-agent config abstraction","D","KILL","AGENTS.md open standard (Linux Foundation, 60K+ projects)","—","No — proposing a rival fractures adoption"],
 ["Local fine-tuning / 'personal AI twin'","D","KILL","Ollama, LM Studio (free); $20/mo inference < $50/mo tune","—","No — consumer TAM collapsed"],
 ["Agent-native terminal (Linux/Windows)","D","CONDITIONAL","Warp $70M (700K users, open-sourced Apr-26)","~$30M+ to compete; effectively a KILL for a solo founder.","No — capital + OS-rendering"],
 ["Generic no-code agent builder","D","KILL","Zapier Agents, n8n $180M ($2.5B val), Make, Lindy","—","No — commoditized category"],
 ["Repository intelligence / AI code review","D","KILL","CodeRabbit $60M, Greptile $25M, Qodo; Copilot native","—","No — consolidation complete"],
 ["Benchmark integrity / synthetic-task gen","D","CONDITIONAL","Labs DIY decontamination; no self-serve SaaS precedent","A 3-5yr regulatory-audit window (AI Act) — hard sell today.","No — research-tier, hard to monetize"],
 ["Prompt-caching optimizer dashboard","D","KILL","Anthropic native cache dashboard","—","No — UI arbitrage; no data lock-in"],
 ["Thinking-time / effort budget optimizer","D","KILL","Anthropic native API knobs (Effort/Extended Thinking)","—","No — a pass-through wrapper"],
 ["Capital-heavy frontier (chips/robotics/biotech)","D","KILL","2-5yr, $5M+, teams","—","No — not a solo/non-technical play"],
]

def build():
    wb = Workbook(); ws = wb.active; ws.title = "VALIDATED MAP"
    # summary banner rows
    n_kill = sum(1 for r in R if r[2]=="KILL"); n_cond = sum(1 for r in R if r[2]=="CONDITIONAL"); n_surv = sum(1 for r in R if r[2]=="SURVIVE")
    ws.append([f"VALIDATED IDEA MAP — {DATE}  |  {len(R)} ideas web-researched (fresh eyes, evidence-only)"])
    ws.append([f"CLEAN SURVIVORS: {n_surv}   |   CONDITIONAL (needs a named unlock): {n_cond}   |   KILLED: {n_kill}"])
    ws.append(["Rubric: SURVIVE only if ALL — (1) real demand (2) gap vs funded incumbents (3) lean/non-technical buildable, no heavy capital/license (4) real moat (data/distribution/community/regulatory, not 'AI wrapper')."])
    ws.append([])
    ws["A1"].font = Font(size=13, bold=True); ws["A2"].font = Font(size=11, bold=True, color="B91C1C")
    ws["A3"].font = Font(italic=True, size=9)
    hdr = 5
    ws.append(COLS)
    for c in range(1, len(COLS)+1):
        cell = ws.cell(row=hdr, column=c); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937"); cell.alignment = Alignment(wrap_text=True, vertical="center")
    order = {"SURVIVE":0,"CONDITIONAL":1,"KILL":2}
    for row in sorted(R, key=lambda x: (order[x[2]], x[1])):
        ws.append(row)
    vfill = {"SURVIVE":"D1FAE5","CONDITIONAL":"FEF3C7","KILL":"FEE2E2"}
    for r in range(hdr+1, hdr+1+len(R)):
        v = ws.cell(row=r, column=3).value
        for c in range(1, len(COLS)+1):
            cell = ws.cell(row=r, column=c); cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill("solid", fgColor=vfill.get(v,"FFFFFF"))
        ws.cell(row=r, column=3).font = Font(bold=True)
    for col,w in zip("ABCDEF",[34,5,12,42,52,30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    out = os.path.join(OUT, f"ai-build-ideas-VALIDATED-{DATE}.xlsx"); wb.save(out)
    cout = os.path.join(OUT, f"ai-build-ideas-VALIDATED-{DATE}.csv")
    with open(cout,"w",newline="",encoding="utf-8") as f:
        w=csv.writer(f); w.writerow(COLS); w.writerows(sorted(R,key=lambda x:(order[x[2]],x[1])))
    return out, cout, n_surv, n_cond, n_kill

if __name__=="__main__":
    x,c,s,co,k = build()
    print("XLSX:",os.path.abspath(x)); print("CSV :",os.path.abspath(c))
    print(f"SURVIVE(clean): {s} | CONDITIONAL: {co} | KILL: {k} | total {s+co+k}")
