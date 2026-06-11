#!/usr/bin/env python3
"""Build the AI-build-ideas deliverable (xlsx + csv) from the 12 scraped channel extracts.
First-principles synthesis: deduped, tiered, scored on defensibility x non-obviousness x money.
Run: python3 scripts/_build-ai-build-ideas.py
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "ideas")
os.makedirs(OUT_DIR, exist_ok=True)
DATE = "2026-06-02"

# Columns
COLS = ["Rank","Tier","Idea","One-line thesis","Why now","Moat type","Moat",
        "Build difficulty","Non-tech-founder friendly?","Source channels","Consensus",
        "Score /5","Verdict","Why defensible — or why it's a trap"]

# Tier A = defensible + real money + non-obvious. B = real consumer/market. C = defensible-ish tooling. D = weak-moat traps (hyped but avoid).
ROWS = [
 # rank, tier, idea, thesis, why-now, moat-type, moat, build, nontech, sources, consensus, score, verdict, note
 [1,"A","Winning-Ad Deconstructor","Monitor competitor ad libraries (Meta/TikTok), auto-score winners, and break each down frame-by-frame (hook timing, pacing, copy, music, CTA) into a repeatable creative spec.","UGC ads are 10x cheaper to make but everyone's blind on WHAT is winning; vision+LLM can now deconstruct a video reliably.","Data (historical ad-performance corpus that compounds)","Strong","Medium (vision + scraping)","YES — buyer is marketers, not devs","Jack Roberts","1","5","BUILD","Real paying market today (Foreplay/Atria/AdEspresso charge $99-300/mo). Your ad corpus compounds into a moat competitors can't backfill. Risk: Meta fights library scraping."],
 [2,"A","Vertical AI Operator (own ONE niche)","Pick one high-value vertical (e-comm returns/refunds, insurance claims triage, clinic front-desk) and own the workflow + ROI benchmark data + integrations — sell the OUTCOME, not 'an AI'.","Build cost is collapsing to zero; business-design value is NOT. Clients can prototype with Claude but can't see which process actually moves the P&L.","Domain data + ROI playbooks + integrations","Strong","Medium (needs domain depth, not heavy eng)","PARTIAL — needs domain expertise or a domain partner","Nate Herk (Devin Kearns)","2","5","BUILD","The one play with a proven $4-5M ARR reference. Moat is the vertical knowledge + case studies, not the model. Pick a vertical with clear $ leverage and documented workflows."],
 [3,"A","AI Sales-Call Sparring Partner","Voice AI that role-plays buyer objections, coaches reps live, scores them on a rubric, and tracks improvement over time — per-vertical objection libraries.","Voice models are now real-time + cheap; sales orgs already track soft KPIs (objection handling, discovery questions).","Objection library per vertical + outcomes data","Medium-Strong","Medium","YES — buyer is sales leaders","Greg Isenberg","1","4","BUILD","Incumbent rep training costs $5K+/rep; async practice is a clear wedge. Integrations (call recording, CRM) get sticky. Niche to one vertical first."],
 [4,"A","Agent Fleet Observability ('FleetView')","The ops layer for companies running 10+ deployed agents: cost, performance, memory/skill drift, A/B rubric scoring, alerting — system-of-record for an agent fleet.","Agents are finally productive enough to deploy at scale; the monitoring/eval layer is missing and painful.","Becoming the system-of-record + eval data","Medium","Medium-High","NO — technical buyer (ML-ops)","Greg Isenberg; Cole Medin","2","4","EXPLORE","Real gap and real pricing leverage, but the window is closing fast and observability players (Datadog, LangSmith) are moving in. Win by niching the eval rubric layer."],
 [5,"A","Dark Factory — gated autonomous shipping","Agents that don't just write code but SHIP it: issue → PR → multi-agent review → tests → canary → monitor → auto-promote or auto-rollback, no human gate. Learns from post-deploy metrics.","Testing+monitoring infra is mature enough to gate deploys without humans; if you ship 10 unattended iterations/day, cost-per-iteration drops 10x.","Private deploy-outcome metrics (compounding)","Strong","Very High (months + prod infra)","NO — deep technical","Cole Medin","1","4","EXPLORE","Highest ceiling + strongest moat (your metrics loop), but the hardest build here and one bad deploy breaks prod. A team play, not a solo non-technical one."],

 [6,"B","Condition-specific Health 'Brain'","One condition only (GERD, migraine, IBS): aggregate bloodwork + biome + diet + specialist input into a single decision layer; monetize meal plans, supplements, specialist matching.","Bloodwork + microbiome APIs are now standardized; patients are fragmented across doctors with no central brain.","Medical-data lock-in + regulatory positioning + LTV","Strong","Medium-High (regulatory)","PARTIAL — needs clinical credibility","Greg Isenberg","1","4","EXPLORE","Very high willingness-to-pay ($200-2K/yr), huge TAM (60M GERD sufferers). Defensible via data + clinical-decision-support positioning. Regulatory complexity is the cost of the moat."],
 [7,"B","Agent-first 'Action App' for a vertical","Mobile app that DOES the task end-to-end (books, cancels, files, disputes) instead of chatting about it — built agent-first, vertical by vertical.","Agent SDKs are mature; incumbents bolt AI onto legacy click-based UX and can't rebuild agent-first.","First-mover vertical UX + workflow depth","Medium","Medium","PARTIAL","Greg Isenberg","1","4","EXPLORE","Greenfield vs legacy productivity apps. Moat is execution + a redesign incumbents can't copy without cannibalizing themselves. Churns hard if automation feels unreliable."],
 [8,"B","IRL Community / anti-loneliness membership","Hybrid online+offline membership that pairs people for real-world experiences (dinners, hobby retreats); revenue from events + membership.","Third-place collapse + loneliness epidemic; proven retreat economics ($90K per 30-person event).","Network effects + offline switching costs","Medium-Strong","Low-Medium (ops-heavy, not tech-heavy)","YES — barely any code","Greg Isenberg","1","4","BUILD","The most durable, least-AI idea here — and the least likely to be eaten by a model release. Ops-intensive, not engineering-intensive. Strong fit for a non-technical operator."],
 [9,"B","AI-native niche media → owned product","Build a genuinely high-quality (NOT slop) audience channel in one vertical; the audience becomes the owned acquisition channel for a downstream product.","Top-1% AI-assisted content beats mid-tier human creators on consistency; algorithm rewards it.","Audience trust (owned, platform-independent)","Medium","Low-Medium","YES","Greg Isenberg","1","3","EXPLORE","Audience is the moat — but only if quality is high; slop dies fast. Treat the channel as a funnel, not the product."],
 [10,"B","Elder-tech for the 65+ market","Apps/devices for older adults (AI hearing/vision aids, memory, social connection); a cash-rich, founder-ignored demographic.","AI unlocks new accessibility; boomers have cash and low founder competition (perceived 'boring').","Demographic trust + regulatory positioning","Medium","Medium","PARTIAL","Greg Isenberg","1","3","EXPLORE","Large TAM, real pain, low competition — but acquisition + retention need different channels. Honest: a slow, trust-first market."],

 [11,"C","Adversarial Agent Testing Harness","Ensemble of agents (fuzzer, security, perf, chaos) that auto-test agent-written code and emit a test suite + vuln report.","Human coverage hits diminishing returns; agents are good at adversarial thinking; CI compute is cheap.","Agent ensemble + feedback loop","Medium","Medium-High","NO — technical","Cole Medin","1","4","EXPLORE","Real leverage as vibe-coded volume explodes. Moderate moat; differentiate via bug-bounty integration + coverage trend data."],
 [12,"C","'Dark-code' QA & validation for production","Toolkit+service to validate AI-generated systems before prod: success-criteria templates, edge-case testing, audit of code nobody actually read.","Teams ship orders of magnitude more unreviewed AI code; traditional QA doesn't scale to it.","Testing library + checklist benchmark","Medium","Medium","PARTIAL","Nate Herk","1","4","EXPLORE","Genuine emerging gap ('dark code'). Moat softens over ~18mo as validation gets baked into the models — move fast or niche to a regulated vertical."],
 [13,"C","Agent Auto-Validation Loop","Drop-in CI layer: separate fresh-context reviewer agents + rule engine (build/test/secret/perf gates) + auto-fix dispatch + a KB that learns from past fixes.","You can't scale parallel agents while humans are the validation bottleneck; separate-context review catches ~40% more.","Proprietary fix-outcome KB","Medium","Medium","NO — technical","Cole Medin","1","4","EXPLORE","High leverage, moderate moat via the learning loop. Plugs into existing CI so adoption friction is low."],
 [14,"C","Self-evolving Agent Knowledge Base","Institutional memory agents read+write: decision logs, what-failed, perf gotchas — so the next agent jumps to the solution instead of repeating mistakes.","Karpathy-style LLM knowledge bases are now feasible; teams are frustrated agents repeat each other's mistakes.","Per-codebase proprietary KB","Medium","Medium","NO","Cole Medin; Nate Herk; Jack Roberts","3","3","EXPLORE","Real pain, but infra is commoditized (Supabase+pgvector in 48h). Likely a feature of a bigger platform than a standalone business."],
 [15,"C","Vision module for headless agents","Give CLI/API-only agents the ability to 'see' — read dashboards, QA screenshots, catch broken layouts — as structured JSON per frame.","Agents can code and write but can't see; screenshot feedback is still manual.","Frame-parsing taxonomy + temporal consistency","Weak-Medium","Medium","NO","Jack Roberts","1","3","EXPLORE","Useful wedge, but high risk Claude/Hermes ship native agent-vision within 3-6 months. Only build with a fast vertical wedge."],

 [16,"D","Multi-model cost router / API-spend optimizer","Route each task to the cheapest sufficient model + dashboard the savings.","Model proliferation + real cost spread (Haiku vs Opus = 5x).","Routing heuristics","WEAK","Low-Medium","NO","Greg Isenberg; Nate Herk; Jack Roberts; AI Explained","5+","2","AVOID","THE most-hyped idea across channels — and the weakest moat. OpenRouter + Anthropic ship this natively. Cross-channel consensus is a trap signal here, not a buy signal."],
 [17,"D","Prompt-caching optimizer dashboard","Analytics + tips to maximize Claude prompt-cache hits and cut token spend.","Caching saves ~90% but is opaque; no third-party tools yet.","Analytics + template lib","WEAK","Low","Partial","Nate Herk","2","2","AVOID","Real insight, ~12-month runway, then Anthropic ships a native dashboard and you're done. Fine as a free lead-gen tool, not a business."],
 [18,"D","Thinking-time budget optimizer","Auto-tune model 'effort'/thinking time vs task difficulty to cut reasoning spend.","Variable thinking time is now first-class across frontier models.","Classifier on task difficulty","WEAK","Low","NO","Nate Herk; AI Explained","2","2","AVOID","Native-feature risk in <3 months. A setting, not a startup."],
 [19,"D","Cross-agent config abstraction","Normalize CLAUDE.md / agents.md / .codex so projects port between agent frameworks with no context loss.","3+ agent frameworks now exist; switching is painful.","Spec + integrations","WEAK","Low","NO","Nate Herk","1","2","AVOID","Doomed by the agent vendors shipping their own normalization. Market still crystallizing."],
 [20,"D","Benchmark aggregator / eval consolidator","One dashboard that runs your model across 20+ public benchmarks and returns a capability card.","No single source of truth; orgs re-run evals independently.","Aggregation only","WEAK","Medium","NO","AI Explained","1","1","AVOID","Aggregating public data + public evals = no defensibility. Benchmark owners would rather own the platform. Cost center, not product."],
 [21,"D","Model drift / honesty canary for prod","Daily canary queries that fingerprint a model version and alert when an API provider silently changes behavior.","Providers ship new model versions monthly; teams can't detect silent behavior shifts.","Canary dataset","Weak-Medium","Low-Medium","NO","AI Explained","1","2","EXPLORE","More defensible than the rest of Tier D (the canary dataset is yours), but 'who pays for a warning?' is unsolved. Validate willingness-to-pay before building."],
 [22,"D","Local fine-tuning / 'personal AI twin' platform","One-click local fine-tune of an open model on your data/values; sovereignty + no API lock-in.","Local inference (Llama/Mistral/Qwen) is viable; LoRA is cheap.","Brand trust on quality","WEAK","Medium","Partial","David Ondrej","2","2","AVOID","Slim margins (local infra), tiny market (privacy cognoscenti), open-source algos. A narrative, not a moat. together.ai/HF already ship the pipeline."],
 [23,"D","Agent-native terminal (Linux/Windows)","A modern multi-agent terminal UI (Cmux-style) for the 95% of agentic devs stuck on Tmux outside macOS.","Devs run 10-20 parallel agents for hours; Tmux UX is 19 years old.","Workflow integration","WEAK","High (OS-level rendering)","NO","David Ondrej","1","2","AVOID","Small TAM (~$1.5M), terminal is a commodity, rendering is hard. Love from devs, no money."],
 [24,"D","Agent skill marketplace w/ creator payouts","Versioned, shareable agent 'skills' with usage analytics + royalty splits to skill authors.","Users hand-write skills from scratch; zero reuse; creator-economy angle.","Marketplace network effects","Weak-Medium","Medium-High","NO","Greg Isenberg","2","2","AVOID (for now)","Pure marketplace dynamics — dead until 10K+ active users exist on a platform you don't control. Gate entry on proven demand; the agent vendors may own this layer."],
]

def build_xlsx():
    wb = Workbook()
    # ---- README sheet ----
    ws0 = wb.active; ws0.title = "README"
    readme = [
        ["NakshIQ — AI 'What to Build' Idea Mine", ""],
        ["Generated", DATE],
        ["Method", "Scraped latest 10 videos x 6 Tier-1 YouTube channels (60 videos) via yt-dlp; 12 Haiku agents extracted ideas first-principles (blank slate, NO existing-business mapping); deduped + scored here."],
        ["Channels", "Greg Isenberg, Nate Herk, Cole Medin, Jack Roberts, AI Explained, David Ondrej"],
        ["Raw extracts", "data/ideas/_raw/*.md  |  Transcripts: .scrapes/youtube/yt-<id>/"],
        ["", ""],
        ["HOW TO READ THIS", ""],
        ["Tier A", "Defensible + real money + non-obvious. The 'sexy AND buildable' shortlist."],
        ["Tier B", "Real consumer/market plays (less AI-infra, more durable — least likely to be eaten by a model release)."],
        ["Tier C", "Defensible-ish dev tooling. Real, but technical buyer + closing moat windows."],
        ["Tier D", "HYPED BUT WEAK-MOAT TRAPS. Listed so you can recognise and AVOID them. The most-pitched ideas live here."],
        ["", ""],
        ["KEY FINDING", "The most-pitched idea across channels (cost router / caching) has the WEAKEST moat — every agent flagged 'Anthropic ships it native'. Consensus != defensibility."],
        ["KEY FINDING", "The defensible ideas cluster on DATA moats (ad corpus, vertical domain data, health data) and AUDIENCE/NETWORK moats — NOT engineering."],
        ["Verdict legend", "BUILD = worth a real validation sprint now | EXPLORE = real but caveated | AVOID = looks sexy, isn't defensible"],
        ["Score", "0-5 on (non-obviousness x moat x money on the table). My opinion, not the YouTuber's."],
    ]
    for r in readme:
        ws0.append(r)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 120
    ws0["A1"].font = Font(size=14, bold=True)
    for i in range(1, len(readme)+1):
        ws0.cell(row=i, column=1).font = Font(bold=True)
        ws0.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Ideas sheet ----
    ws = wb.create_sheet("IDEAS (ranked)")
    ws.append(COLS)
    head_fill = PatternFill("solid", fgColor="1F2937")
    for c in range(1, len(COLS)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = head_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    tier_fill = {"A":"D1FAE5","B":"DBEAFE","C":"FEF3C7","D":"FEE2E2"}
    for row in ROWS:
        ws.append(row)
    for r in range(2, len(ROWS)+2):
        tier = ws.cell(row=r, column=2).value
        fill = PatternFill("solid", fgColor=tier_fill.get(tier,"FFFFFF"))
        for c in range(1, len(COLS)+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill
    widths = [6,5,26,46,42,26,10,18,20,22,11,9,11,52]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i) if i<=26 else "A"+chr(64+i-26)].width = w
    ws.freeze_panes = "A2"

    out = os.path.join(OUT_DIR, f"ai-build-ideas-{DATE}.xlsx")
    wb.save(out)
    return out

def build_csv():
    out = os.path.join(OUT_DIR, f"ai-build-ideas-{DATE}.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COLS)
        w.writerows(ROWS)
    return out

if __name__ == "__main__":
    x = build_xlsx(); c = build_csv()
    print("XLSX:", os.path.abspath(x))
    print("CSV :", os.path.abspath(c))
    print(f"Ideas: {len(ROWS)} | Tier A: {sum(1 for r in ROWS if r[1]=='A')} | B: {sum(1 for r in ROWS if r[1]=='B')} | C: {sum(1 for r in ROWS if r[1]=='C')} | D(traps): {sum(1 for r in ROWS if r[1]=='D')}")
