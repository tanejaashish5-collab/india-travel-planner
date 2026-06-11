#!/usr/bin/env python3
"""Convert data/festivals/festival-celebration-prompts.csv → .xlsx (formatted).
Bold frozen header, sensible column widths, wrapped text for long prompt cells,
and a colour flag on rows that already have real footage. Run after the .mjs
generator: python3 scripts/_csv-to-xlsx-festival-prompts.py"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "data", "festivals", "festival-celebration-prompts.csv")
OUT = os.path.join(ROOT, "data", "festivals", "festival-celebration-prompts.xlsx")

with open(SRC, newline="", encoding="utf-8") as f:
    rows = list(csv.reader(f))
header, data = rows[0], rows[1:]

wb = Workbook(); ws = wb.active; ws.title = "celebration-prompts"
hdr_fill = PatternFill("solid", fgColor="1F2937")
has_real = PatternFill("solid", fgColor="DCFCE7")  # light green
ws.append(header)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    c.alignment = Alignment(vertical="center")
ws.freeze_panes = "A2"

idx = {h: i for i, h in enumerate(header)}
for r in data:
    ws.append(r)
    if r[idx["has_real_footage"]] == "yes":
        ws.cell(row=ws.max_row, column=idx["has_real_footage"] + 1).fill = has_real

widths = {"festival_name": 34, "destination": 18, "state": 16, "month_name": 11,
          "footage_family": 18, "celebration": 60, "full_video_prompt": 110,
          "negative_prompt": 40, "page_url_en": 44, "reference_image_url": 30,
          "festival_slug": 30, "approximate_date": 18, "prompt_grounding": 14}
for h in header:
    col = get_column_letter(idx[h] + 1)
    ws.column_dimensions[col].width = widths.get(h, 12)
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=header[c.column - 1] in ("celebration", "full_video_prompt", "negative_prompt"))

wb.save(OUT)
print(f"Wrote {ws.max_row - 1} rows → {os.path.relpath(OUT, ROOT)}")
