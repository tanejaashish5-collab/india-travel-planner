#!/usr/bin/env python3
"""
One-shot builder: writes the trek-fill-2026-05-27 AI image-prompt
spreadsheet (xlsx + csv) for the 21 destinations added in commits
5b8763c8 / 2c3be205 that don't yet have a card image at
apps/web/public/images/destinations/<slug>.jpg.

Re-run after editing the PROMPTS list below.
    python3 scripts/_build-trek-fill-image-prompts.py
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "data" / "ai-image-prompts"
STEM = "trek-fill-2026-05-27"

LOCKED_SPEC = (
    "Photoreal 16:9 editorial travel photograph, 1376×768. Documentary style — "
    "35mm film grain, slightly muted natural colours, soft contrast, deep depth of field. "
    "Soft early-morning or late-afternoon light unless otherwise specified. "
    "No people posing, no text, no logos, no captions. Calm, credible, magazine cover quality. "
    "Composition leaves the lower-third slightly darker / less detailed so an editorial title overlay can sit there cleanly."
)

# (slug, name, state, cluster, iconic_view_prompt)
PROMPTS = [
    ("adi-kailash", "Adi Kailash", "Uttarakhand", "Garhwal — sacred peaks",
     "The 5,945m snow-faced pyramid of Adi Kailash (Chhota Kailash) rising over the high-altitude Jolingkong meadow and the sacred Gauri Kund (Parvati Tal) reflecting the peak at dawn, Kumaon's far-eastern Himalayas near the Tibet border. Cold blue sky, low golden light raking the snow, brown alpine grass in the foreground, a few stone cairns and prayer flags weathered by wind. Empty landscape, no people. Documentary editorial."),
    ("kinner-kailash", "Kinner Kailash", "Himachal Pradesh", "Garhwal — sacred peaks",
     "The 6,050m granite massif of Kinner Kailash above Kalpa village, Himachal Pradesh, with the distinctive 79-foot vertical Shivling-shaped rock formation on its face catching first sunrise. Apple orchards and slate-roofed Kinnauri houses in the mid-ground, Satluj valley falling away. Crisp morning air, pink-gold light on the snow, deep shadow on the valley. Editorial documentary."),
    ("shrikhand-mahadev", "Shrikhand Mahadev", "Himachal Pradesh", "Garhwal — sacred peaks",
     "The 5,155m Shrikhand Mahadev peak in Kullu, Himachal Pradesh, with its summit Shivling-shaped rock spire piercing a clear high-altitude sky. Foreground of scree, late-summer alpine flowers, a single weathered trail-marker stone. Hard sunlight on rock, deep cool shadow on the snow patches. Severe, magazine-cover composition. No people."),
    ("manimahesh-kailash", "Manimahesh Kailash", "Himachal Pradesh", "Garhwal — sacred peaks",
     "The sacred peak Manimahesh Kailash (5,653m) in Bharmour, Chamba, mirrored in the still surface of Manimahesh Lake at first light. Stone offerings and a faint dirt path along the lake edge in the foreground, brown alpine grass, a few unmanned brass trishuls planted in cairns. Cold high-altitude blue sky, golden-pink alpenglow on the snow face. Reverent, empty, documentary."),
    ("mansarovar-kailash", "Mansarovar Kailash", "Uttarakhand", "Garhwal — sacred peaks",
     "The unmistakable black pyramidal mass of Mt Kailash rising beyond the turquoise expanse of Lake Mansarovar, seen from the Indian-side approach at Lipulekh / Nabhidhang. Vast trans-Himalayan plateau, ochre and grey ground, faint snow streaks on Kailash's south face. Cold thin-air light, crystal clarity, no people, no structures. Pilgrimage stillness."),
    ("vishnuprayag", "Vishnuprayag", "Uttarakhand", "Garhwal — Panch Prayag",
     "The confluence of the Alaknanda and Dhauliganga rivers in a steep grey-stone gorge near Joshimath, milky glacial water of the Alaknanda meeting the darker Dhauliganga, foaming where they merge. Iron suspension footbridge crossing high above, deodar slopes climbing on either side, faint Garhwal village rooftops on the far bank. Soft overcast monsoon-edge light, mist on the gorge walls. Documentary, no figures."),
    ("nandaprayag", "Nandaprayag", "Uttarakhand", "Garhwal — Panch Prayag",
     "The confluence of the Alaknanda and Nandakini rivers below Nandaprayag town, Garhwal — wide turquoise sangam, broad pebble shoals, stone ghats descending from the slope, a small white shikhara temple on the bluff above. Pine-clad ridges in soft hazy afternoon light, smoke from one chulha rising thin from the town. Editorial, no posing figures."),
    ("karnaprayag", "Karnaprayag", "Uttarakhand", "Garhwal — Panch Prayag",
     "The confluence of the milky Alaknanda and emerald Pindar rivers at Karnaprayag, Garhwal — two distinct colours visibly merging into one stream, terraced stone ghats, an old red-and-white temple cluster on the spur above, hillside town of slate roofs climbing the slope behind. Late-afternoon side light, deep depth of field. Documentary calm."),

    ("palani", "Palani", "Tamil Nadu", "South — Arupadai Veedu + Tirumala",
     "The Murugan Arulmigu Dhandayuthapani temple crowning the Palani hilltop, Tamil Nadu, with its gold-leaf vimana shrine catching pre-sunset light against a dry-blue sky. The famous 693 stone steps zigzagging down the rock face in the mid-ground. Coconut palms and a dusty plains town spreading away into haze below. Editorial, no pilgrims posing. Warm muted south-Indian palette."),
    ("tiruparankundram", "Tiruparankundram", "Tamil Nadu", "South — Arupadai Veedu + Tirumala",
     "The Arupadai Veedu Tiruparankundram Murugan temple carved directly into the base of the monolithic granite Parankundram hill near Madurai, Tamil Nadu, white-and-red striped Dravidian gopuram set against the bare rock face. A small temple tank in the foreground with stone steps and a single granite mandapam. Late-afternoon warm light raking the rock. Editorial, empty courtyard."),
    ("tiruchendur", "Tiruchendur", "Tamil Nadu", "South — Arupadai Veedu + Tirumala",
     "The Tiruchendur Subramaniya Swamy temple right on the shore of the Bay of Bengal, Thoothukudi district, Tamil Nadu — multi-tiered Dravidian gopuram in white and bright temple-red, ocean waves breaking on the rocks immediately beside the temple wall, wet black sand. Soft pre-monsoon overcast sea light, salty mist. Empty beach, no figures. Magazine cover quality."),
    ("swamimalai", "Swamimalai", "Tamil Nadu", "South — Arupadai Veedu + Tirumala",
     "The Swamimalai Murugan temple on its low artificial hill in the Cauvery delta near Thanjavur, Tamil Nadu — squat ochre-and-white Dravidian gopuram visible above lush coconut palms, the wide green Cauvery (Kaveri) river with paddy fields and a stone ghat in the foreground. Golden-hour late light, soft river haze. Documentary, no posing figures."),
    ("tiruttani", "Tiruttani", "Tamil Nadu", "South — Arupadai Veedu + Tirumala",
     "The Tiruttani Murugan temple on its 365-step hilltop near Chennai, Tamil Nadu, weathered Dravidian gopuram rising over scrub-jungle slopes and granite boulders. Wide aerial-style view of the rocky hill plus the dusty plains and one curving road far below. Warm late-afternoon haze, soft muted dry-country palette. No people posing."),
    ("pazhamudircholai", "Pazhamudircholai", "Tamil Nadu", "South — Arupadai Veedu + Tirumala",
     "The Pazhamudircholai Murugan temple set inside dense semi-evergreen forest on the Alagar Hills near Madurai, Tamil Nadu — a small ornate gopuram half-hidden by tall trees, sunlight breaking through canopy onto the temple stones, a clear forest stream and natural rock pool (Noopura Gangai) in the foreground. Cool green forest light, deep documentary depth of field. No figures."),
    ("tirumala", "Tirumala", "Andhra Pradesh", "South — Arupadai Veedu + Tirumala",
     "The Sri Venkateswara temple complex on the Tirumala hills above Tirupati, Andhra Pradesh — view from the seventh hill at first light, gold gopuram glinting through morning mist, the white temple roofline, surrounding dense jungle of the Seshachalam range falling away in seven undulating ridges. Cool dawn blue-grey light turning gold on the gopuram. Reverent scale, no posing figures."),

    ("bhoramdeo", "Bhoramdeo", "Chhattisgarh", "Central — Chhattisgarh + Jharkhand",
     "The 11th-century Bhoramdeo Shaiva temple near Kawardha, Chhattisgarh — black-grey sandstone shikhara densely carved with Khajuraho-style erotic and divine figures, set against forested Maikal hills. Open temple courtyard with mossy stone, one weathered nandi bull facing the sanctum. Soft side light, the carvings deeply shadowed and three-dimensional. Documentary editorial, no figures."),
    ("chitrakote-falls", "Chitrakote Falls", "Chhattisgarh", "Central — Chhattisgarh + Jharkhand",
     "The horseshoe-shaped Chitrakote Falls on the Indravati river, Bastar district, Chhattisgarh — India's widest waterfall in full monsoon flood, thick brown-ochre water thundering over a 95-foot semicircular cliff into a deep pool, dense sal and teak forest closing the rim. Spray haze catching late-afternoon light, a faint rainbow at the base. Empty viewpoint, no figures. Magazine-cover scale."),
    ("betla", "Betla", "Jharkhand", "Central — Chhattisgarh + Jharkhand",
     "A wide sal forest clearing at Betla National Park, Palamau, Jharkhand — late-afternoon golden light slanting through tall sal trunks, a chital herd grazing in the mid-distance, the ruined stone walls of Palamu Fort half-swallowed by jungle on the right. Layered foliage, deep depth of field, no people. Editorial wildlife documentary palette."),

    ("jampui-hills", "Jampui Hills", "Tripura", "Northeast",
     "The Betlingchhip viewpoint in Tripura's Jampui Hills at sunrise — endless folded green ridges of Mizoram-style hills disappearing into a low cloud-sea below, pine and orange-orchard slopes in the foreground, a single tin-roofed Lushai village clinging to the next ridge. Cold dawn light turning warm on the eastern ranges. Reverent silence, no figures."),
    ("kiphire", "Kiphire", "Nagaland", "Northeast",
     "Mount Saramati's twin-peak 3,841m massif rising over the dense subtropical jungle of southern Nagaland near Kiphire — viewed from a ridge clearing with thatched Sangtam tribal village rooftops in the foreground, banana and bamboo groves, mist filling the valleys below. Soft post-monsoon golden light, rich green palette. No posing figures. Documentary editorial."),
    ("ziro", "Ziro", "Arunachal Pradesh", "Northeast",
     "The Ziro valley in Arunachal Pradesh — patchwork of bright green Apatani rice-and-fish paddies stitched together by raised earth bunds, low pine-clad hills wrapping the bowl, traditional Apatani bamboo-and-thatch houses clustered in one village on the far edge. Soft monsoon-edge afternoon light, layered haze on the hills, a single farmer's bamboo footbridge in the foreground. No posing figures. Editorial UNESCO-tentative-list scale."),
]

SAVE_DIR_REL = "apps/web/public/images/destinations/"
DIM = "1376×768"
FORMAT = "JPG, ≤350 KB, sRGB"

ROWS = [
    {
        "n": i + 1,
        "slug": slug,
        "name": name,
        "state": state,
        "cluster": cluster,
        "filename": f"{slug}.jpg",
        "save_path": f"{SAVE_DIR_REL}{slug}.jpg",
        "dimensions": DIM,
        "format": FORMAT,
        "status": "pending",
        "full_prompt": f"{LOCKED_SPEC}\n\n{view}",
        "view_only": view,
    }
    for i, (slug, name, state, cluster, view) in enumerate(PROMPTS)
]


def write_csv():
    path = OUT_DIR / f"{STEM}.csv"
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(["#", "slug", "name", "state", "cluster", "filename", "save_path",
                    "dimensions", "format", "status", "full_prompt"])
        for r in ROWS:
            w.writerow([r["n"], r["slug"], r["name"], r["state"], r["cluster"],
                        r["filename"], r["save_path"], r["dimensions"], r["format"],
                        r["status"], r["full_prompt"]])
    return path


def write_xlsx():
    wb = Workbook()

    # ─────────────── README ───────────────
    readme = wb.active
    readme.title = "README"
    readme.sheet_properties.tabColor = "E55642"
    readme.column_dimensions["A"].width = 110

    sections = [
        ("NakshIQ — Trek-fill AI image prompts (2026-05-27)", True),
        ("", False),
        (f"Destinations to render: {len(ROWS)}", False),
        ("Source: 25 destinations added by commits 5b8763c8 / 2c3be205 (trek-fill batch). 4 already had card images, 21 need rendering.", False),
        ("", False),
        ("▎ Output spec (every image)", True),
        (f"  • Dimensions: {DIM}  (16:9 landscape — matches existing 505 dest cards)", False),
        (f"  • Format: {FORMAT}", False),
        ("  • Visual language: photoreal documentary editorial. 35mm grain, muted natural colour, soft golden-hour or early-morning light, no people posing, no text, no logos, no captions. Lower-third slightly darker for overlay parity.", False),
        ("", False),
        ("▎ Where to save the rendered .jpg", True),
        (f"  {SAVE_DIR_REL}<slug>.jpg", False),
        ("  e.g.  apps/web/public/images/destinations/adi-kailash.jpg", False),
        ("  (one file per row in the PROMPTS sheet — the `filename` column is canonical)", False),
        ("", False),
        ("▎ Naming rule", True),
        ("  Exact lowercase slug match — same `id` the destination row has in Supabase. The card components (destination-thumb.tsx, treks-content.tsx) hardcode this convention; any case/typo and the card 404s and renders blank.", False),
        ("", False),
        ("▎ Universal locked spec — already prepended to every full_prompt in the PROMPTS sheet:", True),
        (LOCKED_SPEC, False),
        ("", False),
        ("▎ Workflow", True),
        ("  1. PROMPTS sheet → copy `full_prompt` into the AI image tool (Nano Banana / Imagen / Co-work / Midjourney — same brand-pack across all).", False),
        ("  2. Render at 1376×768. If the tool only outputs square, render at 1536×1536 and crop to 1376×768 centred on the iconic landmark.", False),
        ("  3. Compress to ≤350 KB JPG (existing cards average 200–300 KB).", False),
        ("  4. Save to the `save_path` column.", False),
        ("  5. Mark `status` → `done` and commit when all 21 are in.", False),
        ("", False),
        ("▎ Commit message when shipping", True),
        ("  feat(web,assets): AI-generated card images for 21 new trek destinations", False),
        ("", False),
        ("▎ Post-ship verification", True),
        ("  Visit /en/treks — every trek card under the new entries should render its hero strip without a grey placeholder.", False),
    ]
    for line, bold in sections:
        row = readme.append([line])
        cell = readme.cell(row=readme.max_row, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            cell.font = Font(bold=True, color="E55642" if line.startswith("NakshIQ") else "1A1A1A",
                             size=14 if line.startswith("NakshIQ") else 11)
    # Bump row heights for the wrapped lines
    for i in range(1, readme.max_row + 1):
        readme.row_dimensions[i].height = 18

    # ─────────────── PROMPTS ───────────────
    sheet = wb.create_sheet("PROMPTS")
    headers = ["#", "slug", "name", "state", "cluster", "filename", "save_path",
               "dimensions", "format", "status", "full_prompt"]
    widths = [4, 22, 22, 18, 28, 26, 50, 12, 18, 12, 90]
    sheet.append(headers)
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = sheet.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 22
    sheet.freeze_panes = "A2"

    for r in ROWS:
        sheet.append([r["n"], r["slug"], r["name"], r["state"], r["cluster"],
                      r["filename"], r["save_path"], r["dimensions"], r["format"],
                      r["status"], r["full_prompt"]])
    for i in range(2, sheet.max_row + 1):
        sheet.row_dimensions[i].height = 110
        for col in range(1, len(headers) + 1):
            sheet.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    dv = DataValidation(type="list", formula1='"pending,drafted,rendered,done,skip"', allow_blank=False)
    sheet.add_data_validation(dv)
    dv.add(f"J2:J{sheet.max_row}")

    # ─────────────── VIEW-ONLY ───────────────
    view = wb.create_sheet("VIEW-ONLY")
    view.append(["slug", "name", "iconic_view_prompt (no locked spec)"])
    view.column_dimensions["A"].width = 22
    view.column_dimensions["B"].width = 22
    view.column_dimensions["C"].width = 100
    for col in range(1, 4):
        c = view.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
    view.freeze_panes = "A2"
    for r in ROWS:
        view.append([r["slug"], r["name"], r["view_only"]])
    for i in range(2, view.max_row + 1):
        view.row_dimensions[i].height = 80
        for col in range(1, 4):
            view.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    path = OUT_DIR / f"{STEM}.xlsx"
    wb.save(path)
    return path


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = write_csv()
    xlsx_path = write_xlsx()
    print(f"✓ wrote {csv_path.relative_to(ROOT)}")
    print(f"✓ wrote {xlsx_path.relative_to(ROOT)} (README + PROMPTS + VIEW-ONLY)")
    print(f"  rows: {len(ROWS)}")


if __name__ == "__main__":
    main()
